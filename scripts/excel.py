from openpyxl import load_workbook, Workbook
import datetime

excel_file = 'data/Reddit-Story-Generator.xlsx'

date = datetime.datetime.now().strftime("%m/%d/%Y")

wb = load_workbook(excel_file)
ws = wb.active

def store_data(postid, title, subreddit, url):
    ws.append([date, postid, title, subreddit, url])
    wb.save(excel_file)

def get_row_index(postid):
    for row_index, row in enumerate(ws.rows):
        if row[1].value == postid:
            return row_index + 1
    return None

def store_story(postid, title, story, caption):
    index = get_row_index(postid)
    ws[f'F{index}'] = title
    ws[f'G{index}'] = story
    ws[f'J{index}'] = caption
    wb.save(excel_file)

def store_content_info(postid, voice, mc_video, final_video):
    index = get_row_index(postid)
    ws[f'H{index}'] = voice
    ws[f'I{index}'] = mc_video
    ws[f'K{index}'] = final_video
    wb.save(excel_file)

def store_schedule_info(postid, date, time):
    index = get_row_index(postid)
    ws[f'M{index}'] = date
    ws[f'N{index}'] = time
    wb.save(excel_file)
